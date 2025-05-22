import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import re
from openpyxl.utils.exceptions import InvalidFileException

# Шлях до Excel-файлу
EXCEL_FILE = "shops.xlsx"

# Список стовпців (позиційно прив’язані до B:S)
COLUMNS = ["С1", "05с", "ЧС", "лC", "Б1", "05б", "ЧБ", "лБ", "1Ф", "05ф", "1Ж", "05ж", "кр", "житній", "пироги",
           "багет", "Іваніні", "Повернення"]

# Словник для зіставлення скорочених назв із повними
SHOP_ALIASES = {
    "калина": "Калина",
    "сяйво": "Сяйво",
    "оленка": "Оленка",
    "преміум": "Преміум Ківерці",
    "жабка": "Муравище(Жабка)",
    "ж": "Муравище(Жабка)",
    "дачне": "Дачне Лілія",
    "конякіна": "Заб. Конякіна",
    "північка": "Північка",
    "варшавка": "Варшавка",
    "кравчука": "Заб. Кравчука",
    "фермерський": "Фермерський",
    "соборності28": "Заб. Соборності 28",
    "соборності11в": "Заб. Соборності 11В",
    "тамтам": "Там-Там",
    "галя липинського": "Галя балувана Липинського",
    "липинського": "Заб. Липинського",
    "тамтам експрес": "Там-Там Експрес",
    "овочі-фрукти": "Овочі-Фрукти",
    "відродження": "Заб. Відродження 1В",
    "галя дубенська": "Галя балувана Дубенська",
    "волі": "Заб. Волі",
    "рованці": "Заб. Рованці",
    "підземелля": "Підземелля",
    "мясна хата": "Мясна Хата Луцьк",
    "галя гпз": "Галя балувана ГПЗ",
    "гпз48": "ГПЗ48",
    "гуд маркет": "ГУД МАРКЕТ",
    "розмай": "Розмай",
    "забіяка львівська": "Забіяка Львівська",
    "розмай львівська": "Розмай львівська",
    "забіяка забороль": "Забіяка Забороль",
    "забіяка ковельська": "Забіяка Ковельська",
    "винниченка": "Заб. Винниченка",
    "грушевського": "Заб. Грушевського",
    "франка": "Франка",
    "вараш": "Вараш",
    "ківерці пекарня": "Ківерці пекарня"
}


# Функція для пошуку повної назви магазину за скороченням
def find_shop_name(short_name: str) -> str:
    short_name_lower = short_name.lower().strip()
    for alias, full_name in SHOP_ALIASES.items():
        if short_name_lower == alias.lower():
            return full_name
    return None


# Функція для парсингу введеного тексту
def parse_input(text: str) -> tuple:
    try:
        # Спробуємо новий формат (наприклад, "калина 100с")
        short_pattern = r"([^\s][\w\s()-]+?)\s+((?:\d+[а-яА-Я/]+)(?:\s+\d+[а-яА-Я/]+)*)"
        short_match = re.match(short_pattern, text.strip(), re.UNICODE)

        if short_match:
            shop_short_name = short_match.group(1).strip()
            items = short_match.group(2).strip()

            # Знаходимо повну назву магазину
            shop_name = find_shop_name(shop_short_name)
            if not shop_name:
                return None, None, f"Магазин '{shop_short_name}' не знайдено. Доступні скорочення: {', '.join(SHOP_ALIASES.keys())}"

            # Парсимо елементи (наприклад, "100с")
            data = {}
            # Розбиваємо на окремі пари "кількість+тип"
            item_pattern = r"(\d+)([а-яА-Я/]+)"
            item_matches = re.findall(item_pattern, items, re.UNICODE)

            if not item_matches:
                return None, None, "Невірний формат хліба. Приклад: 'калина 100с'"

            unrecognized = []
            for count, bread_type in item_matches:
                count = int(count)
                bread_type = bread_type.lower()
                if bread_type == "с":
                    data["С1"] = data.get("С1", 0) + count  # 100 кг сірого
                elif bread_type == "б":
                    data["Б1"] = data.get("Б1", 0) + count  # 1 кг білого
                elif bread_type == "/с":
                    data["05с"] = data.get("05с", 0) + count  # Половинка сірого
                elif bread_type == "/б":
                    data["05б"] = data.get("05б", 0) + count  # Половинка білого
                elif bread_type == "чс":
                    data["ЧС"] = data.get("ЧС", 0) + count
                elif bread_type == "лс":
                    data["лC"] = data.get("лC", 0) + count
                elif bread_type == "чб":
                    data["ЧБ"] = data.get("ЧБ", 0) + count
                elif bread_type == "лб":
                    data["лБ"] = data.get("лБ", 0) + count
                elif bread_type == "ф":
                    data["1Ф"] = data.get("1Ф", 0) + count
                elif bread_type == "/ф":
                    data["05ф"] = data.get("05ф", 0) + count
                elif bread_type == "ж":
                    data["1Ж"] = data.get("1Ж", 0) + count
                elif bread_type == "/ж":
                    data["05ж"] = data.get("05ж", 0) + count
                elif bread_type == "к" or bread_type == "кр":
                    data["кр"] = data.get("кр", 0) + count
                elif bread_type == "житній":
                    data["житній"] = data.get("житній", 0) + count
                elif bread_type == "пироги":
                    data["пироги"] = data.get("пироги", 0) + count
                elif bread_type == "багет":
                    data["багет"] = data.get("багет", 0) + count
                elif bread_type == "іваніні":
                    data["Іваніні"] = data.get("Іваніні", 0) + count
                else:
                    unrecognized.append(f"{count}{bread_type}")

            if not data:
                return None, None, f"Не вдалося розпізнати типи хліба. Нерозпізнані значення: {', '.join(unrecognized)}. Доступні типи: с, б, /с, /б, чс, лс, чб, лб, ф, /ф, ж, /ж, к, житній, пироги, багет, іваніні"
            if unrecognized:
                result_label.config(text=f"Попередження: Нерозпізнані значення: {', '.join(unrecognized)}")
            return shop_name, data, None

        # Якщо формат зі скороченнями не підійшов, перевіряємо старий формат (наприклад, "Калина С1=1шт 05с=3шт")
        old_pattern = r"([^\s][\w\s()-]+?)\s*((?:(?:[0-9а-яА-Я]+=[0-9]+шт|[0-9а-яА-Я]+=[0-9а-яА-Я/ ]+)\s*)+)"
        old_match = re.match(old_pattern, text.strip(), re.UNICODE)

        if old_match:
            shop_short_name = old_match.group(1).strip()
            items = old_match.group(2)

            # Знаходимо повну назву магазину
            shop_name = find_shop_name(shop_short_name)
            if not shop_name:
                return None, None, f"Магазин '{shop_short_name}' не знайдено. Доступні скорочення: {', '.join(SHOP_ALIASES.keys())}"

            # Парсимо елементи (наприклад, "С1=1шт")
            item_pattern = r"([0-9а-яА-Я]+)=([0-9а-яА-Я/ ]+?)(?=\s*[0-9а-яА-Я]+=|$)"
            item_matches = re.findall(item_pattern, items, re.UNICODE)

            data = {}
            for key, value in item_matches:
                key_lower = key.lower()
                matched_col = None
                for col in COLUMNS:
                    if key_lower == col.lower():
                        matched_col = col
                        break
                if matched_col:
                    if matched_col == "Повернення":
                        data[matched_col] = value.strip()
                    else:
                        if not value.replace("шт", "").strip().isdigit():
                            return None, None, f"Невірне значення для {key}: має бути числове (наприклад, 1шт)"
                        data[matched_col] = int(value.replace("шт", "").strip())
                else:
                    return None, None, f"Стовпець {key} не підтримується. Доступні: {', '.join(COLUMNS)}"
            return shop_name, data, None

        return None, None, "Невірний формат введення. Приклади: 'калина 100с' або 'Калина С1=1шт 05с=3шт'"
    except Exception as e:
        return None, None, f"Помилка парсингу: {str(e)}"


# Функція для оновлення Excel-файлу
def update_excel(shop_name: str, data: dict, file_path: str) -> str:
    try:
        # Завантажуємо Excel-файл
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Прив’язуємося до позицій стовпців (B:S)
        headers = {}
        for idx, col in enumerate(COLUMNS, start=2):  # Починаємо з 2 (стовпець B)
            headers[col] = idx

        shop_col = 1  # Стовпець A для назв магазинів

        # Знаходимо рядок із магазином
        row_idx = None
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=shop_col).value
            if cell_value == shop_name:
                row_idx = row
                break

        if row_idx is None:
            # Якщо магазин не знайдено, додаємо новий рядок
            row_idx = ws.max_row + 1
            while ws.cell(row=row_idx - 1, column=1).value and "Всього" in str(
                    ws.cell(row=row_idx - 1, column=1).value):
                row_idx -= 1
            ws.cell(row=row_idx, column=shop_col).value = shop_name

        # Оновлюємо значення
        for key, value in data.items():
            if key in headers:
                ws.cell(row=row_idx, column=headers[key]).value = value
            else:
                return f"Помилка: Стовпець {key} не знайдено в таблиці"

        # Зберігаємо файл
        wb.save(file_path)
        wb.close()
        return f"Дані для {shop_name} успішно оновлено!"
    except PermissionError:
        return "Помилка: Файл Excel заблоковано або відкрито в іншій програмі"
    except InvalidFileException:
        return "Помилка: Вибрано некоректний файл Excel"
    except Exception as e:
        return f"Помилка при оновленні Excel: {str(e)}"


# Функція для відображення таблиці
def display_table(file_path: str):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # Очищаємо таблицю
        for row in table.get_children():
            table.delete(row)

        # Перевіряємо, чи є дані
        if ws.max_row < 2:
            result_label.config(text="Файл Excel порожній або не містить даних")
            wb.close()
            return

        # Додаємо дані, ігноруючи підсумки
        data_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and "Всього" not in str(row[0]) and "НА ПОРІЗКУ" not in str(row[0]):
                data_rows.append([v if v is not None else "" for v in row[:len(COLUMNS) + 1]])

        if not data_rows:
            result_label.config(text="Не знайдено даних для відображення (перевірте формат файлу)")
            wb.close()
            return

        # Додаємо дані в таблицю
        for row in data_rows:
            table.insert("", "end", values=row)

        # Оновлюємо підсумки
        totals = ["Всього"]
        for col_idx in range(1, len(COLUMNS)):
            if COLUMNS[col_idx - 1] != "Повернення":
                total = sum(float(row[col_idx]) for row in data_rows if
                            row[col_idx] and str(row[col_idx]).replace(".", "").isdigit())
                totals.append(total)
            else:
                totals.append("")
        total_label.config(text=" | ".join(f"{COLUMNS[i - 1]}: {totals[i]}" for i in range(1, len(totals)) if
                                           totals[i] and COLUMNS[i - 1] != "Повернення"))

        wb.close()
    except PermissionError:
        messagebox.showerror("Помилка", "Файл Excel заблоковано або відкрито в іншій програмі")
    except InvalidFileException:
        messagebox.showerror("Помилка", "Вибрано некоректний файл Excel")
    except Exception as e:
        messagebox.showerror("Помилка", f"Не вдалося завантажити таблицю: {str(e)}")


# Функція для вибору Excel-файлу
def select_file():
    global EXCEL_FILE
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if file_path:
        EXCEL_FILE = file_path
        file_label.config(text=f"Вибрано: {file_path}")
        display_table(file_path)


# Функція для редагування клітинки в таблиці
def edit_cell(event):
    try:
        item = table.selection()[0]
        col = table.identify_column(event.x)
        col_idx = int(col.replace("#", "")) - 1
        if col_idx == 0:  # Не дозволяємо редагувати назву магазину
            return

        # Отримуємо значення
        row_values = table.item(item, "values")
        shop_name = row_values[0]
        column_name = COLUMNS[col_idx - 1]

        # Створюємо поле для редагування
        edit_entry = tk.Entry(table_frame)
        edit_entry.insert(0, row_values[col_idx])
        edit_entry.place(x=event.x_root - root.winfo_rootx(), y=event.y_root - root.winfo_rooty())

        def save_edit(event=None):
            new_value = edit_entry.get().strip()
            data = {}
            if column_name == "Повернення":
                data[column_name] = new_value
            else:
                if new_value.replace("шт", "").strip().isdigit():
                    data[column_name] = int(new_value.replace("шт", "").strip())
                else:
                    messagebox.showerror("Помилка", f"Значення для {column_name} має бути числовим (наприклад, 1шт)")
                    edit_entry.destroy()
                    return

            result = update_excel(shop_name, data, EXCEL_FILE)
            if "успішно" in result:
                display_table(EXCEL_FILE)
            else:
                messagebox.showerror("Помилка", result)
            edit_entry.destroy()

        edit_entry.bind("<Return>", save_edit)
        edit_entry.bind("<FocusOut>", save_edit)
        edit_entry.focus_set()
    except Exception as e:
        messagebox.showerror("Помилка", f"Не вдалося редагувати: {str(e)}")


# Функція для обробки введення
def process_input():
    text = input_field.get("1.0", tk.END).strip()
    shop_name, data, error = parse_input(text)

    if error:
        result_label.config(text=error)
        return

    if shop_name and data:
        result = update_excel(shop_name, data, EXCEL_FILE)
        result_label.config(text=result)
        if "успішно" in result:
            display_table(EXCEL_FILE)  # Оновлюємо таблицю
    else:
        result_label.config(text="Невірний формат. Приклади: 'калина 100с' або 'Калина С1=1шт 05с=3шт'")


# Створення головного вікна
root = tk.Tk()
root.title("Excel Bot для пекарні")
root.geometry("1000x700")

# Поле для вибору файлу
file_button = tk.Button(root, text="Вибрати Excel-файл", command=select_file)
file_button.pack(pady=5)

file_label = tk.Label(root, text="Excel-файл не вибрано")
file_label.pack()

# Поле для введення тексту
tk.Label(root, text="Введіть дані (наприклад, 'калина 100с' або 'Калина С1=1шт 05с=3шт'):").pack(pady=5)
input_field = tk.Text(root, height=3, width=50)
input_field.pack()

# Кнопка для обробки
process_button = tk.Button(root, text="Оновити Excel", command=process_input)
process_button.pack(pady=10)

# Мітка для результату
result_label = tk.Label(root, text="")
result_label.pack(pady=5)

# Мітка для підсумків
total_label = tk.Label(root, text="", wraplength=900)
total_label.pack(pady=5)

# Таблиця для відображення даних
table_frame = tk.Frame(root)
table_frame.pack(fill="both", expand=True, padx=10, pady=10)

table = ttk.Treeview(table_frame, columns=["Назва магазину"] + COLUMNS, show="headings")
for col in ["Назва магазину"] + COLUMNS:
    table.heading(col, text=col)
    table.column(col, width=80)
table.pack(fill="both", expand=True)

# Скролбар для таблиці
scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=table.yview)
table.configure(yscroll=scrollbar.set)
scrollbar.pack(side="right", fill="y")

# Прив'язка редагування до подвійного кліку
table.bind("<Double-1>", edit_cell)

# Запуск головного циклу
root.mainloop()